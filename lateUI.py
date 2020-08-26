import tkinter as tk
import tkinter.messagebox
import xlrd
from tkinter import filedialog
import pyodbc
import datetime
import openpyxl
from Logger_format import config_log

class ui_frame:
    def __init__(self):
        main_log.info(" the Init function called ")
        self.window = tk.Tk()
        # background grey color

        self.bg = "#d2d4dc"
        self.bg_title = "#4c516d"
        self.window.config(bg=self.bg)
        self.frame_a = tk.Frame(bg=self.bg)
        self.frame_title = tk.Frame(bg=self.bg, padx=30)
        self.frame_b = tk.Frame()
        self.frame_b.config(bg=self.bg, padx=20)
        self.photo = tk.PhotoImage(file=r"kk.png")
        self.cross_box = tk.PhotoImage(file=r"crossButton.png")
        main_log.info(" the frame A is set with titles ")
        main_log.info(" the frame B with the test names and descriptions is added")

        self.frameCheck = tk.Frame()
        self.frameCheck.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck2 = tk.Frame()
        self.frameCheck2.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck3 = tk.Frame()
        self.frameCheck3.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck4 = tk.Frame()
        self.frameCheck4.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck5 = tk.Frame()
        self.frameCheck5.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck6 = tk.Frame()
        self.frameCheck6.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck7 = tk.Frame()
        self.frameCheck7.config(bg=self.bg, padx=15, pady=3)
        self.frameCheck8 = tk.Frame()
        self.frameCheck8.config(bg=self.bg, padx=15, pady=3)

        self.framebutton = tk.Frame()
        self.framebutton.config(bg=self.bg, pady=7)

        self.meta_check_1 = tk.IntVar()
        self.meta_check_2 = tk.IntVar()
        self.integrity_check_1 = tk.IntVar()
        self.integrity_check_2 = tk.IntVar()
        self.quality_check = tk.IntVar()
        self.completeness_check = tk.IntVar()
        self.cleansing_check = tk.IntVar()
        self.transformation_check = tk.IntVar()
        self.xml_data_load_check = tk.IntVar()
        self.stored_proc_check = tk.IntVar()
        main_log.info(" UI is completed defining")

        try:
            self.conn = pyodbc.connect("Driver={ODBC Driver 17 for SQL Server};"
                                       "Server=CDC2-L-F1TGQEH;"
                                       "Database=test;"
                                       "Trusted_Connection=yes;"
                                       )
            self.cursor = self.conn.cursor()
            main_log.info("Connection established with object - " + self.cursor)
        except Exception as e:
            main_log.error(" connection Failed due to - ")


    def calling_frame_a(self):
        # Heading-blue
        logo = tk.PhotoImage(file=r"new.png")
        label_a = tk.Label(master=self.frame_a, image=logo)
        label_a.image = logo
        label_a.pack(pady=(0, 20))
        label_a.config(font=("Calibri", "24", "bold"))

        # SubHeading-grey
        label_b = tk.Label(self.frame_title, text="Test Types", fg="white", bg=self.bg_title, anchor="w")
        label_b.config(font=("Calibri", "12", "bold"))
        label_b.grid(row=3, column=1, pady=15, sticky="w", ipadx=46, ipady=5)
        label_b = tk.Label(self.frame_title, text="Test Objective", fg="white", bg=self.bg_title, width=72, anchor="w")
        label_b.config(font=("Calibri", "12", "bold"))
        label_b.grid(row=3, column=2, padx=20, ipady=5, ipadx=89, sticky="w")
        label_b = tk.Label(self.frame_title, text="Applicable Test Phase ", fg="white", bg=self.bg_title, anchor="w")
        label_b.config(font=("Calibri", "12", "bold"))
        label_b.grid(row=3, column=3, sticky="w", ipady=5, padx=10, ipadx=13)
        label_a.pack()

    def calling_frame_b(self):

        #  ROW ONE - Metadata Test
        label_b = tk.Label(self.frame_b, text="MetaData Test", fg="white", bg="#0057e7", activebackground='#0057e7')
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=1, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate target tables are built as per the source table schema definition/data definition model",
                           fg="black", bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=1, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW TWO - DataIntegrity test
        label_b = tk.Label(self.frame_b, text="Data Integrity Test", fg="white", bg="#0057e7",
                           activebackground='#0057e7')
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=5, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that the referential integrity between the tables are established accurately",
                           fg="black", bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=5, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW THREE - Data Quality
        label_b = tk.Label(self.frame_b, text="Data Quality Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=6, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that the content/records of the table are as per the data type and constraint definition of the element",
                           fg="black", bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=6, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW F0UR - Data completeness Test
        label_b = tk.Label(self.frame_b, text="Data Completeness Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=7, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b, text="To validate the accuracy data load from source to target tables",
                           fg="black",
                           bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=7, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW FIVE - Data Cleansing test
        label_b = tk.Label(self.frame_b, text="Data Cleansing Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=8, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that the elements of target table against the data cleansing rules",
                           fg="black",
                           bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=8, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW SIX - Data transformation test
        label_b = tk.Label(self.frame_b, text="Data Transformation Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=9, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that data transformed into IDS is as per the business transoformation logic",
                           fg="black",
                           bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=9, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW SEVEN - hop2 -xml Data load Test
        label_b = tk.Label(self.frame_b, text="Hop2 - XML Data Load Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=10, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that the xml data in IDS Staging is loaded appropriately into IDS table as per the XSD definition.",
                           fg="black",
                           bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=10, column=4, padx=10, pady=10, ipady=7, sticky="e,w")

        # ROW EIGHT - Hop3 Stored proc test
        label_b = tk.Label(self.frame_b, text="Hop3 - Stored Proc Test", fg="white", bg="#0057e7")
        label_b.config(font=("Calibri", "11", "bold"))
        label_b.grid(row=11, column=3, padx=10, pady=10, ipady=7, sticky="e,w")
        label_b = tk.Label(self.frame_b,
                           text="To validate that the xml generated by the Hop3 Stored Proc is accurate and has all requried IDS data based on extract query",
                           fg="black",
                           bg="white", anchor="w")
        label_b.config(font=("Calibri", "11"))
        label_b.grid(row=11, column=4, padx=10, pady=10, ipady=7, sticky="e,w")
        main_log.info("labels with test case names are populated ")

# checkbox frames
    def checkbox_frame1(self):

        label_b = tk.Label(self.frameCheck, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")

        check_b = tk.Checkbutton(self.frameCheck, bg=self.bg, variable=self.meta_check_1)
        check_b.grid(row=0, column=2, sticky="e,w")

        label_b = tk.Label(self.frameCheck, text="   Hop2", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck, bg=self.bg, variable=self.meta_check_2)
        check_b.grid(row=0, column=4, padx=2, ipady=2)

    def checkbox_frame2(self):
        label_b = tk.Label(self.frameCheck2, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")

        check_b = tk.Checkbutton(self.frameCheck2, bg=self.bg, variable=self.integrity_check_1)
        check_b.grid(row=0, column=2, sticky="e,w")

        label_b = tk.Label(self.frameCheck2, text="   Hop2", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck2, bg=self.bg, variable=self.integrity_check_2)
        check_b.grid(row=0, column=4, padx=2, ipady=2)

    def checkbox_frame3(self):
        label_b = tk.Label(self.frameCheck3, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")
        label_image_cross = tk.Label(self.frameCheck3, text="\u274e", bg=self.bg)
        label_image_cross.grid(row=0, column=2)
        label_b = tk.Label(self.frameCheck3, text="     Hop2 ", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck3, bg=self.bg, variable=self.quality_check)
        check_b.grid(row=0, column=4, sticky="e,w", ipady=2)

    def checkbox_frame4(self):
        label_b = tk.Label(self.frameCheck4, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck4, bg=self.bg, variable=self.completeness_check)
        check_b.grid(row=0, column=2, sticky="e,w")
        label_b = tk.Label(self.frameCheck4, text="   Hop2", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        label_image_cross = tk.Label(self.frameCheck4, text="\u274e", bg=self.bg)
        label_image_cross.grid(row=0, column=4, padx=4.5, ipady=5)

    def checkbox_frame5(self):
        label_b = tk.Label(self.frameCheck5, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")
        label_image_cross = tk.Label(self.frameCheck5, text="\u274e", bg=self.bg)
        label_image_cross.grid(row=0, column=2)
        label_b = tk.Label(self.frameCheck5, text="     Hop2 ", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck5, bg=self.bg, variable=self.cleansing_check)
        check_b.grid(row=0, column=4, sticky="e,w", pady=2, ipady=2)

    def checkbox_frame6(self):
        label_b = tk.Label(self.frameCheck6, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")
        label_image_cross = tk.Label(self.frameCheck6, text="\u274e", bg=self.bg)
        label_image_cross.grid(row=0, column=2)
        label_b = tk.Label(self.frameCheck6, text="     Hop2 ", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck6, bg=self.bg, variable=self.transformation_check)
        check_b.grid(row=0, column=4, sticky="e,w", pady=2)

    def checkbox_frame7(self):
        label_b = tk.Label(self.frameCheck7, text="Hop1", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1, sticky="e,w")
        label_image_cross = tk.Label(self.frameCheck7, text="\u274e", bg=self.bg)
        label_image_cross.grid(row=0, column=2)
        label_b = tk.Label(self.frameCheck7, text="     Hop2 ", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=3, sticky="e,w")
        check_b = tk.Checkbutton(self.frameCheck7, bg=self.bg, variable=self.xml_data_load_check)
        check_b.grid(row=0, column=4, sticky="e,w", pady=2)

    def checkbox_frame8(self):
        label_b = tk.Label(self.frameCheck8, text="          Hop3", bg=self.bg)
        label_b.config(font=("", "10", "bold"))
        label_b.grid(row=0, column=1)
        check_b = tk.Checkbutton(self.frameCheck8, bg=self.bg, variable=self.stored_proc_check)
        check_b.grid(row=0, column=4, pady=2)

# Buttons Frames
    def buttons_frame(self):
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.meta_data_test)
        label_b_button.grid(row=1, column=1, padx=10, pady=2)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.data_integrity_test)
        label_b_button.grid(row=2, column=1, padx=10, pady=11)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.data_quality_test)
        label_b_button.grid(row=3, column=1, padx=10, pady=1)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.data_completeness_test)
        label_b_button.grid(row=4, column=1, padx=10, pady=11)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.data_cleansing_test)
        label_b_button.grid(row=5, column=1, padx=10, pady=3)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg)
        label_b_button.grid(row=6, column=1, padx=10, pady=8)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.hop2_xml_data_load_test)
        label_b_button.grid(row=7, column=1, padx=10, pady=4.5)
        label_b_button = tk.Button(self.framebutton, text='Click Me !', image=self.photo, relief="flat", bg=self.bg,
                                   command=self.hop3_stored_proc_test)
        label_b_button.grid(row=8, column=1, padx=10, pady=5)

        main_log.info("The frames with checkboxes and play button are populated")

# meta data test
    def meta_data_test(self):
        meta_data_log = config_log("metaDataTest", "metaDataTest.log")
        meta_data_log.info("Meta data test Called ")
        if self.meta_check_1.get() == self.meta_check_2.get() == 0:
            meta_data_log.error(" Please select Applicable Test Phase for execution")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            meta_data_log.info(" Checkbox input is received")
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            meta_data_log.info("File name received is - " + filename)
            #workbook
            wb = xlrd.open_workbook(filename)
            print(len(wb.sheet_names()))
            print(wb.sheet_names())

            if self.meta_check_1.get() == self.meta_check_2.get() == 1:

                if "Hop1_MetaData_Input" in wb.sheet_names() and 'Hop2_MetaData_Input' in wb.sheet_names():
                    meta_data_log.info(" Both check box inputs selected ")
                    # hop1 sheet
                    sheet = wb.sheet_by_name('Hop1_MetaData_Input')
                    print(sheet.nrows)
                    sheet_col = sheet.ncols
                    # hop2 sheet
                    sheet2 = wb.sheet_by_name('Hop2_MetaData_Input')
                    print(sheet2.nrows)
                    sheet2_col = sheet2.ncols
                    if sheet.nrows >= 2:
                        if sheet2.nrows >= 2:
                            meta_data_log.info(" The sheets have more than one record. GOOD to proceed.")
                            # Starting to check the cells

                            for col in range(sheet_col):
                                try:
                                    col_type = sheet.cell_type(1, col)
                                    print(col_type)
                                    if col_type == 0:
                                        tk.messagebox.showinfo("error", "Metadata-Hop1 Input sheet have empty values")
                                        meta_data_log.error(" Input sheet with empty cells ")
                                    else:
                                        continue
                                        # functionalities to be added
                                except Exception as e:
                                    col_type = 0
                            print("sheet2")
                            for col2 in range(sheet2_col):
                                try:
                                    col2_type = sheet2.cell_type(1, col2)
                                    print(col2_type)
                                    if col2_type == 0:
                                        tk.messagebox.showinfo("error", "Metadata-Hop2 Input sheet has a empty values ")
                                        meta_data_log.error(" Input sheet with empty cells ")
                                    else:
                                        continue
                                        # functionalities to be added
                                except :
                                    pass

                        else:
                            tk.messagebox.showinfo("Important Message",
                                                   "Metadata-Hop2 Input sheet does not have any input to run the test. ")
                            meta_data_log.error(" Metadata-Hop2 Input sheet does not have any input to run the test.")
                    else:
                        tk.messagebox.showinfo("Important Message",
                                               "Metadata-Hop1 Input sheet does not have any input to run the test. ")
                        meta_data_log.error("Metadata-Hop1 Input sheet does not have any input to run the test.")
                else:
                    print("not present")
                    tk.messagebox.showinfo("Important Message", "Input File doesn't have the required Sheets"
                                                                "\n File Location : " + filename)
                    meta_data_log.error("Input File doesn't have the required Sheets - File Locatoin" + filename)
                    breakpoint()

            elif self.meta_check_1.get() == 1:

                meta_data_log.info("Hop 1 check box selected")
                if "Hop1_MetaData_Input" in wb.sheet_names():
                    sheet_Hop1 = wb.sheet_by_name('Hop1_MetaData_Input')

                    if sheet_Hop1.nrows >= 2:
                        meta_data_log.info(" The input sheets have more than 1 record ")

                        for col in range(sheet_Hop1.ncols):
                            try:
                                c_h1_type = sheet_Hop1.cell_type(1, col)
                                print(c_h1_type)
                                if c_h1_type == 0:
                                    tk.messagebox.showinfo("Error", "Metadata-Hop1 Input sheet have empty values")
                                    meta_data_log.error("Metadata-Hop1 Input sheet have empty values")
                                else:
                                    # functionalites to be added
                                    continue
                            except Exception:
                                c_h1_type = 0
                    else:
                        tk.messagebox.showinfo("Important Message",
                                               "Metadata-Hop1 Input sheet does not have any input to run the test. ")
                        meta_data_log.error("Metadata-Hop1 Input sheet does not have any input to run the test.")
                else:
                    print(" hop 1 sheet not present")
                    tk.messagebox.showinfo("Important Message", "Input File doesn't have the required Sheets"
                                                                "\n File Location : " + filename)
                    meta_data_log.error("Input File doesn't have the required Sheets - Filename :" + filename)
                    breakpoint()


            else:
                meta_data_log.info("Hop 2 check box is selected ")
                if "Hop2_MetaData_Input" in wb.sheet_names():

                    sheet_Hop2 = wb.sheet_by_name('Hop2_MetaData_Input')
                    if sheet_Hop2.nrows >= 2:
                        meta_data_log.info("The input sheet have more than 1 record")
                        for col_h2 in range(sheet_Hop2.ncols):
                            try:
                                c2_h2_type = sheet_Hop2.cell_type(1, col_h2)
                                print(c2_h2_type)
                                if c2_h2_type == 0:
                                    tk.messagebox.showinfo("Error", "Metadata-Hop2 Input sheet has a empty values ")
                                    meta_data_log.error("Metadata-Hop2 Input sheet has a empty values")
                                else:
                                    continue
                                    # fucntionalities to be added
                            except:
                                pass
                    else:
                        tk.messagebox.showinfo("Important Message",
                                               " Metadata-Hop2 Input sheet does not have any input to run the test. ")
                        meta_data_log.error("Metadata-Hop2 Input sheet does not have any input to run the test.")
                else:
                    print("not present")
                    tk.messagebox.showinfo("Important Message", "Input File doesn't have the required Sheets"
                                                                "\n File Location : " + filename)
                    meta_data_log.error("Input File doesn't have the required Sheets - File Locatoin - " + filename)
                    breakpoint()

    def data_integrity_test(self):
        data_integrity_log = config_log("data_integrity_test","data_integrity_test.log")

        data_integrity_log.info("Data integrity test Started ")
        if self.integrity_check_1.get() == self.integrity_check_2.get() == 0:

            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
            data_integrity_log.error(" Please select the applicable test phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            data_integrity_log.info(wb +" File location " + filename)

            if self.integrity_check_1.get() == self.integrity_check_2.get() == 1:
                data_integrity_log.info(" Both check boxes selected ")
                if "Hop1_DataIntegrity_Input" in wb.sheet_names() and 'Hop2_DataIntegrity_Input' in wb.sheet_names():
                    data_integrity_log.info("Both Hop1 and Hop2 Sheets present")
                    sheet = wb.sheet_by_name('Hop1_DataIntegrity_Input')
                    sheet_col = sheet.ncols
                    sheet2 = wb.sheet_by_name('Hop2_DataIntegrity_Input')
                    sheet2_col = sheet2.ncols
                    if sheet.nrows >= 2:
                        if sheet2.nrows >= 2:
                            data_integrity_log.info("The sheet has more than 1 Record")
                            # Starting to check the cells

                            for column in range(sheet_col - 2):
                                try:
                                    column_type = sheet.cell_type(1, column)
                                    if column_type == 0:
                                        tk.messagebox.showinfo("error",
                                                               "DataIntegrity-Hop1 Input sheet have empty values")
                                        data_integrity_log.error("DataIntegrity-Hop1 Input sheet have empty values")
                                    else:
                                        continue
                                except:
                                    column_type = 0
                            data_integrity_log.info(" Data Integrity hop2 validation starts")
                            for col_h2 in range(sheet2_col - 2):
                                try:
                                    c2_type = sheet2.cell_type(1, col_h2)
                                    print(c2_type)
                                    if c2_type == 0:
                                        tk.messagebox.showinfo("error",
                                                               "DataIntegrity-Hop2 Input sheet has a empty values ")
                                        data_integrity_log.error("DataIntegrity-Hop2 Input sheet has a empty values")
                                    else:
                                        continue
                                except:
                                    pass

                        else:
                            tk.messagebox.showinfo("Important Message",
                                                   "DataIntegrity-Hop2 Input sheet does not have any input to run the test. ")
                            data_integrity_log.error("DataIntegrity-Hop2 Input sheet does not have any input to run the test. ")
                    else:
                        tk.messagebox.showinfo("Important Message",
                                               "DataIntegrity-Hop1 Input sheet does not have any input to run the test. ")
                        data_integrity_log.error("DataIntegrity-Hop1 Input sheet does not have any input to run the test.")
                else:
                    print("not present")
                    tk.messagebox.showinfo("Important Message", "Input File doesn't have the required Sheets"
                                                                "\n File Location : " + filename)
                    data_integrity_log.error("Input File doesn't have the required Sheets - File location "+filename)
                    breakpoint()

            elif self.integrity_check_1.get() == 1:
                data_integrity_log.info("Hop1 Check box is selected")
                sheet_Hop1 = wb.sheet_by_name('Hop1_DataIntegrity_Input')
                if sheet_Hop1.nrows >= 2:
                    data_integrity_log.info("The sheets have more than 1 record ")
                    for col in range(sheet_Hop1.ncols - 2):
                        try:
                            col_type = sheet_Hop1.cell_type(1, col)
                            print(col_type)
                            if col_type == 0:
                                tk.messagebox.showinfo("error", "DataIntegrity-Hop1 Input sheet have empty values")
                                data_integrity_log.error("DataIntegrity-Hop1 Input sheet have empty values")
                            else:
                                continue
                        except:
                            col_type = 0
                else:
                    tk.messagebox.showinfo("Important Message",
                                           "DataIntegrity-Hop1 Input sheet does not have any input to run the test. ")
                    data_integrity_log.error("DataIntegrity-Hop1 Input sheet does not have any input to run the test. ")
            else:
                data_integrity_log.info(" Hop 2 check box is selected ")
                sheet_Hop2 = wb.sheet_by_name('Hop2_DataIntegrity_Input')
                if sheet_Hop2.nrows >= 2:
                    data_integrity_log.info(" The sheet have more than 1 Record ")
                    for col_h2 in range(sheet_Hop2.ncols - 2):
                        try:
                            c2_type = sheet_Hop2.cell_type(1, col_h2)
                            print(c2_type)
                            if c2_type == 0:
                                tk.messagebox.showinfo("error", "DataIntegrity-Hop2 Input sheet has a empty values ")
                                data_integrity_log.error("DataIntegrity-Hop2 Input sheet has a empty values ")
                            else:
                                continue
                        except:
                            pass
                else:
                    tk.messagebox.showinfo("Important Message",
                                           " DataIntegrity-Hop2 Input sheet does not have any input to run the test. ")
                    data_integrity_log.error("DataIntegrity-Hop2 Input sheet does not have any input to run the test.")

    def data_quality_test(self):

        data_quality_log = config_log("data_quality_test","data_quality_test.log")
        data_quality_log.info("Data quality Test Started ")
        if self.quality_check.get() == 0:
            data_quality_log.error("Please select the applicable test phase for execution")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            data_quality_log.info(" filename received")

            print(len(wb.sheet_names()))
            print(wb.sheet_names())
            cells_present_flag = 1
            if self.quality_check.get() == 1:
                data_quality_log.info("Hop 1 check box is selected ")
                if 'Hop2_DataQuality_Input' in wb.sheet_names():
                    sheet_hop2 = wb.sheet_by_name('Hop2_DataQuality_Input')

                    wb_open = openpyxl.load_workbook(filename)
                    sh = wb_open["Hop2_DataQuality_Input"]
                    if sheet_hop2.nrows > 2:
                        data_quality_log.info("The sheet has more than 1 record")
                        for row in range(1, sheet_hop2.nrows):
                            for col in range(sheet_hop2.ncols - 2):
                                cells_present_flag = 1
                                try:
                                    cell_value = sheet_hop2.cell_value(row, col)
                                    if cell_value:
                                        # print("present" + r)
                                        print("")
                                    else:
                                        tk.messagebox.showinfo("Error", "DataQuality-Hop2 Input sheet have empty cells")
                                        cells_present_flag = 0
                                        data_quality_log.error("DataQuality-Hop2 Input sheet have empty cells")
                                        break
                                except Exception as e:
                                    print(e)
                            if 1 == cells_present_flag:
                                table_name = sheet_hop2.cell_value(row, 2)
                                dates = []
                                rating = []
                                null_rows = []
                                employee_name = []
                                data_quality_log.info(" select query for table ")
                                self.cursor.execute('select * from {}'.format(table_name))
                                for r_row in self.cursor:

                                    dates.append(r_row[2])
                                    employee_name.append(r_row[3])
                                    rating.append(r_row[4])
                                    # null row check
                                    for cell in r_row:
                                        # print(c)
                                        if cell is None:
                                            null_rows.append(1)
                                            break
                                data_quality_log.info("TABLE DATA ")
                                data_quality_log.info("dates ")
                                data_quality_log.info("rating ")
                                data_quality_log.info("employee_name")

                                # reading the validation check from the column 6
                                target_employee_name = []
                                self.cursor.execute('select * from target_employee')
                                for table_row in self.cursor:
                                    target_employee_name.append(table_row[2])


                                choose_Validation_check = sheet_hop2.cell_value( row , 5)
                                if choose_Validation_check == 'Null Check':
                                    dq_null_check = ""
                                    data_quality_log.info("null Values ")
                                    if len(null_rows) == 0:
                                        dq_null_check = "pass"
                                    else:
                                        dq_null_check = "fail"

                                    data_quality_log.info("Null check  - {}".format(dq_null_check))
                                    sh.cell(row= row+1, column= 8).value = dq_null_check



                                elif choose_Validation_check == 'Data Truncation Check':

                                    # max length of employee_name
                                    ml_source_employee = 0
                                    e_len = 0
                                    dq_trunc_check = ""
                                    for employee in list(employee_name):
                                        if employee is not None:
                                            e_len = len(employee)
                                            if e_len > ml_source_employee:
                                                ml_source_employee = e_len
                                    # max length of target employee_name
                                    ml_target_employee = 0
                                    for target_employee in list(target_employee_name):
                                        if target_employee is not None:
                                            te_len = len(target_employee)
                                            if te_len > ml_target_employee:
                                                ml_target_employee = te_len

                                    print(ml_source_employee, "the max")
                                    print(ml_target_employee, "the target max ")
                                    if ml_source_employee == ml_target_employee:
                                        dq_trunc_check = "pass"
                                    else:
                                        dq_trunc_check = "fail"
                                    sh.cell(row=row + 1, column=8).value = dq_trunc_check
                                    data_quality_log.info(f" trunc Check - {dq_trunc_check}")

                                elif choose_Validation_check == 'Duplicate Check':
                                    print("dup")
                                    table_duplicate_check = ""
                                    target_table_duplicate_check = ""
                                    self.cursor.execute(
                                        "select Employee_name, count(*) from {} group by Employee_name having count(*) > 1".format(
                                            table_name))
                                    table_val = self.cursor.fetchall()
                                    print(table_val)

                                    if len(table_val) == 0:
                                        table_duplicate_check = "pass"
                                    else:
                                        table_duplicate_check = "fail"
                                    print(table_duplicate_check)
                                    # second primary column checking duplicate

                                    self.cursor.execute(
                                        "select employee_level_code, count(*) from {} group by employee_level_code having count(*) > 1".format(
                                            table_name))
                                    table_target_val = self.cursor.fetchall()
                                    print(table_target_val)
                                    if len(table_target_val) == 0:
                                        target_table_duplicate_check = "pass"
                                    else:
                                        target_table_duplicate_check = "fail"
                                    print(target_table_duplicate_check)

                                    if table_duplicate_check == target_table_duplicate_check == "pass":
                                        dq_duplicate_check = "pass"
                                    else:
                                        dq_duplicate_check = "fail"
                                    print("duplicate check ", dq_duplicate_check)
                                    data_quality_log.info("duplicate Check - " +dq_duplicate_check)
                                    sh.cell(row=row + 1, column=8).value = dq_duplicate_check

                                elif choose_Validation_check == "Num Field Format":
                                    print("numm")
                                    print(rating)
                                    dq_num_check = ""
                                    for rate in rating:
                                        if rate:
                                            if type(rate) is int:
                                                dq_num_check = "pass"
                                            else:
                                                dq_num_check = "fail"
                                                break
                                    print("num check ", dq_num_check)
                                    data_quality_log.info(" Num check ")
                                    sh.cell(row=row + 1, column=8).value = dq_num_check

                                elif choose_Validation_check == "Date Field Format":
                                    print("date")
                                    dq_date_check = ""
                                    for date in dates:
                                        if date:
                                            if type(date) is datetime.date:
                                                dq_date_check = "pass"
                                            else:
                                                dq_date_check = "fail"
                                                break
                                    print("date check ")
                                    sh.cell(row=row + 1, column=8).value = dq_date_check
                                    data_quality_log.info("date check ")
                                else:
                                    pass

                            else:
                                print("No record is fully filled ")
                            print("\n")
                        wb_open.save(filename)
                        data_quality_log.info(" Data pass/fail written in excel and saved")
                            # print(null_rows)
                    else:
                        tk.messagebox.showinfo("Error", "Hop2 : input sheet does not have the inputs to run the test")
                        data_quality_log.error("Hop2 : input sheet does not have the inputs to run the test")
                else:
                    tk.messagebox.showinfo("Error", "Input File does not have the required sheet"
                                                    "\n File Location : " + filename)
                    data_quality_log.error("Input File does not have the required sheet - File location " + filename)

    def data_completeness_test(self):
        data_completeness_log = config_log("data_completeness_test","data_completeness_test.log")
        data_completeness_log.info("data completeness is selected")
        if self.completeness_check.get() == 0:
            data_completeness_log.error(" Please select the applicable test phase for execution ")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            print(len(wb.sheet_names()))
            print(wb.sheet_names())
            data_completeness_log.info("Workbook - Filename "+ filename)

            if self.completeness_check.get() == 1:
                data_completeness_log.info("Hop 1 check box is selected ")
                sheet_Hop1 = wb.sheet_by_name('Hop1_DataCompleteness_Input')
                if sheet_Hop1.nrows >= 2:
                    data_completeness_log.info("The sheet has more than 1 Record")
                    for i in range(sheet_Hop1.ncols - 3):
                        try:
                            c = sheet_Hop1.cell_type(1, i)
                            print(c)
                            if c == 0:
                                tk.messagebox.showinfo("error", "DataCompleteness-Hop1 Input sheet have empty values")
                                data_completeness_log.error("DataCompleteness-Hop1 Input sheet have empty values")
                            else:
                                continue
                        except:
                            c = 0
                else:
                    tk.messagebox.showinfo("Important Message",
                                           "DataCompleteness-Hop1 Input sheet does not have any input to run the test. ")
                    data_completeness_log.error("DataCompleteness-Hop1 Input sheet does not have any input to run the test.")

    def data_cleansing_test(self):
        data_cleansing_log = config_log("data_cleansing_test","data_cleansing_test.log")
        data_cleansing_log.info("data cleansing test is selected")
        if self.cleansing_check.get() == 0:
            data_cleansing_log.error("Please select the applicable test phase for execution")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            print(len(wb.sheet_names()))
            print(wb.sheet_names())
            data_cleansing_log.info(" workbook - filename " + filename)

            if self.cleansing_check.get() == 1:

                data_cleansing_log.info("Hop 2 check boxes is selected")
                sheet_Hop2 = wb.sheet_by_name('Hop2_DataCleansing_Input')
                if sheet_Hop2.nrows >= 2:
                    data_cleansing_log.info("The sheet has more than 1 record")
                    for i in range(sheet_Hop2.ncols - 2):
                        try:
                            c = sheet_Hop2.cell_type(1, i)
                            print(c)
                            if c == 0:
                                tk.messagebox.showinfo("error", "DataCleansing-Hop2 Input sheet have empty values")
                                data_cleansing_log.error("DataCleansing-Hop2 Input sheet have empty values")
                            else:
                                continue
                        except:
                            c = 0
                else:
                    tk.messagebox.showinfo("Important Message",
                                           "Hop2_DataCleansing_Input-Hop1 Input sheet does not have any input to run the test. ")
                    data_cleansing_log.error("Hop2_DataCleansing_Input-Hop1 Input sheet does not have any input to run the test. ")

    # def data_transformation_test(self):
    #     if self.transformation_check.get() == 0:
    #
    #         tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
    #     else:
    #         filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
    #         wb = xlrd.open_workbook(filename)
    #         print(len(wb.sheet_names()))
    #         print(wb.sheet_names())
    #         # sheet = wb.sheet_by_index(0)
    #         # print(sheet.nrows)
    #         # print("columns")
    #         # print(sheet.ncols)
    #         # os.system(filename)
    #
    #         if self.cleansing_check.get() == 1:
    #             print("hop 2 is selected with variable " + str(self.integrity_check_1.get()))
    #             sheet_Hop2 = wb.sheet_by_name('Hop2_DataCleansing_Input')
    #             if sheet_Hop2.nrows >= 2:
    #                 print("proceed")
    #                 for i in range(sheet_Hop2.ncols - 2):
    #                     try:
    #                         c = sheet_Hop2.cell_type(1, i)
    #                         print(c)
    #                         if c == 0:
    #                             tk.messagebox.showinfo("error", "DataCleansing-Hop2 Input sheet have empty values")
    #                             print(" empty cell")
    #                         else:
    #                             continue
    #                     except:
    #                         c = 0
    #             else:
    #                 tk.messagebox.showinfo("Important Message",
    #                                        "Hop2_DataCleansing_Input-Hop1 Input sheet does not have any input to run the test. ")
    #
    #
    #

    def hop2_xml_data_load_test(self):
        hop2_xml_data_load_log = config_log("hop2_xml_data_load_test","hop2_xml_data_load_test.log")
        hop2_xml_data_load_log.info("Hop1 xml data load test is called ")
        if self.xml_data_load_check.get() == 0:
            hop2_xml_data_load_log.error(" Please select the applicable test phase for execution")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            print(len(wb.sheet_names()))
            print(wb.sheet_names())
            hop2_xml_data_load_log.info(" workbook - filename " + filename)

            if self.xml_data_load_check.get() == 1:
                print("hop 2 is selected with variable " + str(self.integrity_check_1.get()))
                sheet_Hop2 = wb.sheet_by_name('Hop2 - XML Data Load Test_Input')
                if sheet_Hop2.nrows >= 2:
                    hop2_xml_data_load_log.info("The sheet has more than 1 Record")
                    for i in range(sheet_Hop2.ncols):
                        try:
                            c = sheet_Hop2.cell_type(1, i)
                            print(c)
                            if c == 0:
                                tk.messagebox.showinfo("error", "XML Data Load Test-Hop2 Input sheet have empty values")
                                hop2_xml_data_load_log.error("XML Data Load Test-Hop2 Input sheet have empty values")
                            else:
                                continue
                        except:
                            c = 0
                else:
                    tk.messagebox.showinfo("Important Message",
                                           "Hop2 - XML Data Load Test Input sheet does not have any input to run the test. ")
                    hop2_xml_data_load_log.error("Hop2 - XML Data Load Test Input sheet does not have any input to run the test.")

    def hop3_stored_proc_test(self):
        hop3_stored_proc_log = config_log("hop3_stored_proc_test","hop3_stored_proc_test.log")

        hop3_stored_proc_log.info("hop3 stored proc test is started ")
        if self.stored_proc_check.get() == 0:
            hop3_stored_proc_log.error(" Please select the applicable test phase for execution")
            tk.messagebox.showinfo(" Error ", " Please Select the Applicable Test Phase for execution")
        else:
            filename = filedialog.askopenfilename(initialdir="C:/", title="select file")
            wb = xlrd.open_workbook(filename)
            print(len(wb.sheet_names()))
            print(wb.sheet_names())
            hop3_stored_proc_log.info("Workbook - filename " + filename)

            if self.stored_proc_check.get() == 1:
                hop3_stored_proc_log.info("The check box is selected")
                sheet_Hop3 = wb.sheet_by_name('Hop3_Store_Proc_Input')
                if sheet_Hop3.nrows >= 2:
                    hop3_stored_proc_log.info("The sheet has more than 1 record")
                    for i in range(sheet_Hop3.ncols - 1):
                        try:
                            c = sheet_Hop3.cell_type(1, i)
                            print(c)
                            if c == 0:
                                tk.messagebox.showinfo("error", "Store_Proc-Hop3 Input sheet have empty values")
                                hop3_stored_proc_log.info("Store_Proc-Hop3 Input sheet have empty values")
                            else:
                                continue
                        except:
                            c = 0
                else:
                    tk.messagebox.showinfo("Important Message",
                                           "Store_Proc_Input-Hop3 Input sheet does not have any input to run the test. ")
                    hop3_stored_proc_log.error("Store_Proc_Input-Hop3 Input sheet does not have any input to run the test.")

    def display(self):
        self.frame_a.pack()
        self.frame_title.pack(anchor="nw")
        self.frame_b.pack(side="left", anchor="nw")
        self.framebutton.pack(side="right", anchor="ne")

        self.frameCheck.config(borderwidth=3.5, relief="ridge")
        self.frameCheck.pack(side="top", pady=11, anchor="w")
        self.frameCheck2.config(borderwidth=3.5, relief="ridge")
        self.frameCheck2.pack(pady=7, anchor="w")
        self.frameCheck3.config(borderwidth=3.5, relief="ridge")
        self.frameCheck3.pack(pady=7, anchor="w")
        self.frameCheck4.config(borderwidth=3.5, relief="ridge")
        self.frameCheck4.pack(pady=7, anchor="w")
        self.frameCheck5.config(borderwidth=3.5, relief="ridge")
        self.frameCheck5.pack(pady=7, anchor="w")
        self.frameCheck6.config(borderwidth=3.5, relief="ridge")
        self.frameCheck6.pack(pady=7, anchor="w")
        self.frameCheck7.config(borderwidth=3.5, relief="ridge")
        self.frameCheck7.pack(pady=7, anchor="w")
        self.frameCheck8.config(borderwidth=3.5, relief="ridge")
        self.frameCheck8.pack(pady=7, ipadx=20, anchor="w")

        self.window.geometry("1260x630")
        self.window.resizable(0, 0)
        self.window.mainloop()
        main_log.info(" The window displayed")

if __name__ == '__main__':
    main_log = config_log("main","main_log.log")
    main_log.info("The Program started")

    ui = ui_frame()


    ui.calling_frame_a()
    ui.calling_frame_b()

    ui.checkbox_frame1()
    ui.checkbox_frame2()
    ui.checkbox_frame3()
    ui.checkbox_frame4()
    ui.checkbox_frame5()
    ui.checkbox_frame6()
    ui.checkbox_frame7()
    ui.checkbox_frame8()

    ui.buttons_frame()

    ui.display()


    main_log.info("Program completed ")
